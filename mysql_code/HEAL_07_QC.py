# /* -------------------------------------------------------------------------------- */
# /* Project: HEAL 																	*/
# /* PI: Kira Bradford																*/
# /* Program: HEAL_07_QC																*/
# /* Programmer: Sabrina McCutchan (CDMS)												*/
# /* Date Created: 2024/05/07															*/
# /* Date Last Updated: 2026/02/10													*/
# /* Description:	This program creates a QC report for data contained in MySQL. 		*/				
# /*		1. progress_tracker table 													*/
# /*		2. awards table																*/
# /*		3. Compare appl_ids in MySQL tables 										*/
# /*		4. Metrics by Study  														*/
# /*																					*/
# /* Notes:  																			*/
# /*		- Added an if condition to prevent errors when no observations exist that	*/
# /*		  meet the selection criteria.												*/
# /*																					*/
# /* -------------------------------------------------------------------------------- */

# ----- Boiler Plate Code -----------------------------------------------------*/


# Import Python Modules Here
import os
import pandas as pd
from datetime import date
import subprocess
import sys
from pathlib import Path
import re
import numpy as np
import openpyxl
from openpyxl import load_workbook  
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# TEMPORARY UNTIL INCLUDED in MASTER
# ==============================================================================
# 0. SET GLOBAL MACROS
# ==============================================================================
today = "20260615"
heal_dir = Path(
    r"C:\Users\berman\OneDrive - Research Triangle Institute\Python Environment\HEAL"
)

# Inject into environment variables for child processes (convert Path to string)
os.environ["today"] = today
os.environ["dir"] = str(heal_dir)
# END TEMPORARY UNTIL INCLUDED in MASTER



# ----- SET MACROS -----*/

# ----- 1. Dates ----- */
today = os.environ.get("today")

# today = date.today().strftime("%Y-%m-%d")
print(today)
# ----- 2. Filepaths ----- */
dir = Path(os.environ.get("dir"))
inp = dir / "Input"
out = dir / "Output"
log = dir / "Log"
qc = out / "QC"

# Logging
log_path = os.path.join(log, f"HEAL_07_QC_{today}_log.txt")
with open(log_path, 'w') as f:
    pass  # 'w' mode truncates existing file or creates new blank file
# open(f"{out}/StudyMetrics_{today}_log.txt", 'w').close() #Clears Log before running

def log_out(message):
    with open(f"{log}/HEAL_07_QC_{today}_log.txt", 'a') as f:
        print(message, file=f)

log_out(f"HEAL_07_QC Log Run Date: {today}")

# ----- END Boiler Plate Code -----------------------------------------------------*/





# ==============================================================================
# 0. Environment Setup & Configuration
# ==============================================================================

# Target report path (Changed extension to .xlsx)
report_path = os.path.join(qc, f"QCReport_{today}.xlsx")

# Initialize OpenPyXL Workbook
wb = Workbook()

# Get the active sheet to start writing data
ws = wb.active
ws.title = "QC Report"  # Optional: Give your sheet a custom name
wb.save(f"{qc}/QCReport_{today}.xlsx")



# ==============================================================================
# 0. Environment Setup & Configuration
# ==============================================================================

# Target report path (Changed extension to .xlsx)
report_path = os.path.join(qc, f"QCReport_{today}.xlsx")

# Initialize OpenPyXL Workbook
wb = Workbook()
ws = wb.active
ws.title = "QC Report"

# Style helper configurations
font_title = Font(name="Calibri", size=14, bold=True, color="1B365D")
font_subtitle = Font(name="Calibri", size=12, bold=True)
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11)
fill_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Light blue Stata-like styling

def add_header(text, size=14):
    """Appends styled section header text into a new row."""
    ws.append([]) # Add visual breathing room before header
    current_row = ws.max_row + 1
    cell = ws.cell(row=current_row, column=1, value=text)
    cell.font = font_title if size == 14 else font_subtitle

def add_text(text):
    """Adds a standard row of descriptive text."""
    current_row = ws.max_row + 1
    cell = ws.cell(row=current_row, column=1, value=text)
    cell.font = font_regular

def add_table_from_df(df, title=None):
    """Converts a Pandas DataFrame into styled rows mimicking standard table grid logic."""
    if title:
        current_row = ws.max_row + 1
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = font_bold
    
    if df.empty:
        ws.append(["No records found matching criteria."])
        ws.cell(row=ws.max_row, column=1).font = font_regular
        ws.append([]) # Visual padding
        return

    # Write Header Row
    ws.append(list(df.columns))
    header_row_idx = ws.max_row
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = font_bold
        cell.fill = fill_header

    # Write Data Rows
    for _, row in df.iterrows():
        # OpenPyXL native types handle strings/numbers automatically better than casting everything to str
        ws.append(list(row.values))
        data_row_idx = ws.max_row
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=data_row_idx, column=col_idx).font = font_regular

    ws.append([]) # Visual padding trailing the table

def get_freq_table(series, col_label, value_labels=None):
    """Generates frequency tables replicating Stata's 'tab' output."""
    freq = series.value_counts(dropna=False).to_frame(name="Frequency")
    freq["Percent"] = (series.value_counts(normalize=True) * 100).round(2)
    freq = freq.reset_index().rename(columns={"index": col_label})
    if value_labels:
        freq[col_label] = freq[col_label].map(value_labels).fillna(freq[col_label])
    return freq

add_header("--------------QC Report--------------", size=20)
add_header(f"--------------Run Date: {today} --------------", size=14)
# ==============================================================================
# 1. progress_tracker table
# ==============================================================================
add_header("", size=14)
add_header("--------------1. progress_tracker table--------------", size=14)
add_text(
    "The MYSQL progress_tracker table is automatically updated daily with a pull of fresh data from the Platform MDS."
)

# Read Progress Tracker file
pt_df = pd.read_csv(os.path.join(inp, f"progress_tracker_{today}.csv"))

# -- 1A. Number of records missing appl_id
pt_df["1a_miss_appl_id"] = (
    pt_df["appl_id"].isna() | (pt_df["appl_id"] == 0)
).astype(int)

counts = pt_df["1a_miss_appl_id"].value_counts(dropna=False)
percentages = (pt_df["1a_miss_appl_id"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_1a = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})

tab_1a["Status"] = tab_1a["Status"].map({0: "0. Not Missing", 1: "1. Missing"})

total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_1a["Count"].sum(), 
    "Percent": tab_1a["Percent"].sum().round(1)
}])

tab_1a = pd.concat([tab_1a, total_row], ignore_index=True)

add_table_from_df(tab_1a, "1A. Number of records missing appl_id")

retain_col = pt_df.pop("1a_miss_appl_id")
pt_df.insert(0, "1a_miss_appl_id", retain_col)


# project number evaluation --
pt_df["project_num_str"] = pt_df["project_num"].fillna("").astype(str).str.strip()
pt_df["1b_miss_proj_num"] = (pt_df["project_num_str"] == "").astype(int)

# Replicating Stata's sieve function (count dashes)
pt_df["num_dashes"] = pt_df["project_num_str"].apply(lambda x: x.count("-"))

# Regex check for CTN protocols
pt_df["ctn_flag"] = (
    pt_df["project_num_str"].str.contains(r"^CTN", flags=re.IGNORECASE).astype(int)
)
pt_df["1c_fmt_proj_num"] = (
    (pt_df["num_dashes"] > 1) | (pt_df["ctn_flag"] == 1)
).astype(int)

# -- 1B. Number of records missing project_num

counts = pt_df["1b_miss_proj_num"].value_counts(dropna=False)
percentages = (pt_df["1b_miss_proj_num"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_1b = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})

tab_1b["Status"] = tab_1b["Status"].map({0: "0. Not Missing", 1: "1. Missing"})

total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_1b["Count"].sum(), 
    "Percent": tab_1b["Percent"].sum().round(1)
}])

tab_1b = pd.concat([tab_1b, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_1b, "1B. Number of records missing project_num")


retain_col = pt_df.pop("1b_miss_proj_num")
pt_df.insert(1, "1b_miss_proj_num", retain_col)


# -- 1C. Number of records with bad format for project_num

counts = pt_df["1c_fmt_proj_num"].value_counts(dropna=False)
percentages = (pt_df["1c_fmt_proj_num"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_1c = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})

tab_1c["Status"] = tab_1c["Status"].map({0: "0. Good Format", 1: "1. Bad Format"})

total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_1c["Count"].sum(), 
    "Percent": tab_1c["Percent"].sum().round(1)
}])

tab_1c = pd.concat([tab_1c, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_1c, "1C. Number of records with bad format for project_num")


retain_col = pt_df.pop("1c_fmt_proj_num")
pt_df.insert(2, "1c_fmt_proj_num", retain_col)


# Report 1D 1D. Number of bad project_number formats due to CTN protocol in project_num field
# bad_formats_df = pt_df[pt_df["project_num_badformat"] == 1]
# tab_1d = get_freq_table(
#     bad_formats_df["ctn_flag"],
#     "CTN Protocol",
#     {0: "Non-CTN Bad Format", 1: "CTN Protocol"},
# )

# -- 1D. Number of bad project_number formats due to CTN protocol in project_num field
pt_df["1d_fmt_ctn_flag"] = ((pt_df["1c_fmt_proj_num"] == 1) & (pt_df["ctn_flag"] == 1)).astype(int)
# subset to just bad proj num format
bad_proj_num_df = pt_df[pt_df["1c_fmt_proj_num"] == 1]

counts = bad_proj_num_df["1d_fmt_ctn_flag"].value_counts(dropna=False)
percentages = (bad_proj_num_df["1d_fmt_ctn_flag"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_1d = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})

tab_1d["Status"] = tab_1d["Status"].map({0: "0. Bad Format not due to CTN Flag", 1: "1. Bad Format Due to CTN Flag"})

total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_1d["Count"].sum(), 
    "Percent": tab_1d["Percent"].sum().round(1)
}])

tab_1d = pd.concat([tab_1d, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_1d, "1D. Number of bad project_number formats due to CTN protocol in project_num field")


retain_col = pt_df.pop("1d_fmt_ctn_flag")
pt_df.insert(3, "1d_fmt_ctn_flag", retain_col)




# ==============================================================================
# 3. Compare appl_ids in MySQL tables
# ==============================================================================
add_header("", size=14)
add_header("", size=14)
add_header("--------------3. Compare appl_ids in MySQL tables--------------", size=14)

# -- Awards and reporter --
nih_tables = pd.read_csv(os.path.join(out, f"nihtables_{today}.csv"))
# tab_3a = get_freq_table(nih_tables["merge_reporter_awards"], "Link Status")

# -- 3A. Compare: reporter and awards
nih_tables["3a_merge_reporter_awards"] = nih_tables["merge_reporter_awards"]
counts = nih_tables["3a_merge_reporter_awards"].value_counts(dropna=False)
percentages = (nih_tables["3a_merge_reporter_awards"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_3a = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})


total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_3a["Count"].sum(), 
    "Percent": tab_3a["Percent"].sum().round(1)
}])

tab_3a = pd.concat([tab_3a, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_3a, "3A. Compare: reporter and awards (using merge_reporter_awards from nih_tables)")

retain_col = nih_tables.pop("3a_merge_reporter_awards")
nih_tables.insert(1, "3a_merge_reporter_awards", retain_col)


# -- MySQL vs MDS --
mysql_df = pd.read_csv(os.path.join(out, f"mysql_{today}.csv"))
mysql_df = mysql_df[mysql_df["entity_type"] != "CTN"]

# tab_3c = get_freq_table(mysql_df["merge_awards_mds"], "Link Status MDS/MySQL")

# -- 3C. Compare: MySQL [reporter & awards] and MDS
mysql_df["3c_merge_awards_mds"] = mysql_df["merge_awards_mds"]

counts = mysql_df["3c_merge_awards_mds"].value_counts(dropna=False)
percentages = (mysql_df["3c_merge_awards_mds"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_3c = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})


total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_3c["Count"].sum(), 
    "Percent": tab_3c["Percent"].sum().round(1)
}])

tab_3c = pd.concat([tab_3c, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_3c, "3C. Compare: MySQL [reporter & awards] and MDS (using merge_awards_mds from mysql)")

retain_col = mysql_df.pop("3c_merge_awards_mds")
mysql_df.insert(1, "3c_merge_awards_mds", retain_col)



# ==============================================================================
# 4. Metrics by Study
# ==============================================================================
add_header("", size=14)
add_header("", size=14)
add_header("--------------4. Metrics by Study--------------", size=14)

# -- 4A. Number of studies, by Stewards xstudy_id
study_lookup = pd.read_csv(os.path.join(out, "study_lookup_table.csv"))

# Unique study count summaries
xstudies = (
    study_lookup["xstudy_id"].dropna().drop_duplicates().astype(int).to_frame()
)
summary_stats = pd.DataFrame([{"Statistic": "xstudy_id", "Distinct Count": len(xstudies)}])
add_header("", size=14)
add_table_from_df(summary_stats, "4A. Number of studies, by Stewards xstudy_id")



# -- 4B. Number of studies with/out a HDP ID

hdp_metrics = study_lookup[["xstudy_id", "study_hdp_id"]].drop_duplicates().copy()
hdp_metrics["4b_study_has_hdp"] = (
    hdp_metrics["study_hdp_id"].notna() & (hdp_metrics["study_hdp_id"] != "")
).astype(int)

counts = hdp_metrics["4b_study_has_hdp"].value_counts(dropna=False)
percentages = (hdp_metrics["4b_study_has_hdp"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_4b = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})

tab_4b["Status"] = tab_4b["Status"].map({0: "0. No HDP_ID", 1: "1. HDP_ID"})

total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_4b["Count"].sum(), 
    "Percent": tab_4b["Percent"].sum().round(1)
}])

tab_4b = pd.concat([tab_4b, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_4b, "4B. Number of studies with/out a HDP ID")


retain_col = hdp_metrics.pop("4b_study_has_hdp")
hdp_metrics.insert(0, "4b_study_has_hdp", retain_col)



# ==============================================================================
# 5. Entity Types
# ==============================================================================
add_header("", size=14)
add_header("", size=14)

add_header("--------------5. Entity Types--------------", size=14)
add_header("--------------5A. Entity Type Rules--------------", size=12)
add_text(
    "There are 3 entity types in our data: Study, CTN Protocol & Other. Every record is assigned Study for entity_type by default. "
    "The entity type is changed to CTN or Other if it meets one of the following conditions:"
)
add_text(
    "- replace entity_type=CTN if the project_num field in Platform/progress_tracker table data is a CTN Protocol number"
)
add_text("- replace entity_type=CTN if res_net==CTN")
add_text(
    "- replace entity_type=Other if the project_num field in Platform/progress_tracker table data does not match the default project number format for grants. For example an OT contract like the Stewards would be replaced entity_type='Other'"
)
add_text("- replace entity_type=Other if appl_id==0")

# -- Entity Type reporting --
add_header("--------------5B. Entity Type: Other--------------", size=12)

mysql_reports = pd.read_csv(os.path.join(out, f"mysql_{today}.csv"))

counts = mysql_reports["entity_type"].value_counts(dropna=False)
percentages = (mysql_reports["entity_type"].value_counts(dropna=False, normalize=True) * 100).round(1)

tab_5b = pd.DataFrame({
    "Status": counts.index,
    "Count": counts.values,
    "Percent": percentages.values
})


total_row = pd.DataFrame([{
    "Status": "Total", 
    "Count": tab_5b["Count"].sum(), 
    "Percent": tab_5b["Percent"].sum().round(1)
}])

tab_5b = pd.concat([tab_5b, total_row], ignore_index=True)

add_header("", size=14)
add_table_from_df(tab_5b, "5B. Entity Type")


other_entities = mysql_reports[mysql_reports["entity_type"] == "Other"][
    ["entity_type","appl_id", "proj_num", "hdp_id", "archived"]
]
other_entities["5b_entity_type"] = other_entities["entity_type"]
retain_col = other_entities.pop("5b_entity_type")
other_entities.insert(0, "5b_entity_type", retain_col)



# ==============================================================================
# Save Completed Quality Control Document
wb.save(f"{qc}/QCReport_{today}.xlsx")


# Export Progress Tracker to QC Report
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    pt_df.to_excel(writer, sheet_name='1_progress_tracker', index=False)

nih_tables = nih_tables.drop(columns=["proj_abs"], errors="ignore")    
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    nih_tables.to_excel(writer, sheet_name='3a_nih_tables', index=False)

mysql_df = mysql_df.drop(columns=["proj_abs"], errors="ignore")    
with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    mysql_df.to_excel(writer, sheet_name='3c_mysql', index=False)

with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    study_lookup.to_excel(writer, sheet_name='4a_study_lookup', index=False)

with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    hdp_metrics.to_excel(writer, sheet_name='4b_hdp_metrics', index=False)

with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    other_entities.to_excel(writer, sheet_name='5b_other_entities', index=False)

