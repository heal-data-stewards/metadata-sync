# -------------------------------------------------------------------------------- */
# Project: HEAL 																	*/
# PI: Kira Bradford, Becky Boyles													*/
# Program: HEAL_00_Master															*/
# Programmer: Brian Erman (RTI)		        										*/
# Date Created: 2024/02/29															*/
# Description:	This is the master Stata program for MySQL data processing. It sets	*/
# global macros before calling the following programs:								*/
#		1. Import & merge data														*/
#		2. Generate Research Networks Table											*/
#		3. Generate Study Table														*/
#		4. Generate CTN crosswalk and outputs										*/
#		98. Generate study metrics report											*/
#		99. Generate QC report														*/
#																					*/
#																					*/
# Notes:  
#	2026/03/20 - Stata code converted to Python                                     */
#	2024/07/18 - The programs that generate the Research Networks table and Study	*/
#		Table were swapped in order due to a dependency in the latter.				*/
#	2024/05/28 - This program originally native to the HEAL_Study program tree.	It	*/
#		has	now been split out because it is a necessary first step to all 			*/
#		processing. This program should be run before any other HEAL programs.		*/

# CURRENTLY IMPORTS CSV files exported from MySQL Heal Studies database
# Requires these _today.csv files to exist:
# pi_emails_yyyy-mm-dd.csv
# reporter_yyyy-mm-dd.csv
# progress_tracker_yyyy-mm-dd.csv
# awards_yyyy-mm-dd.csv
# research_networks_yyyy-mm-dd.csv 

# Import Python Modules Here
import os
import pandas as pd
from datetime import date
import subprocess
import sys
from pathlib import Path
import re
import numpy as np
import openpyxl
from openpyxl import load_workbook  
 
# ----- SET MACROS -----*/

# ----- 1. Dates ----- */
today = "2026-03-04"
# today = date.today().strftime("%Y-%m-%d")
# print(today)

# ----- 2. Filepaths ----- */
# dir = Path(r"/rtpnfil03/rtpnfil03_vol4/NPTB2/EDC/Migrated/HEAL") #UNIX
dir = Path(r"\\rtpnfil03\NPTB2\EDC\Migrated\HEAL") #WINDOwS
raw = dir / "Extracts"
der = dir / "Derived"
prog = dir / "Programs"
doc = dir / "Documentation"
temp = dir / "temp"
out = dir / "Output"
qc = out / "QC"
backups = dir / "Backups"

# ----- 3. Variables ----- */
# Variables used to identify studies
stewards_id_vars = [
    "proj_ser_num",
    "subproj_id",
    "proj_num_spl_sfx_code"
]

key_vars = [
    "study_id",
    "xstudy_id_stewards",
    "appl_id",
    "hdp_id",
    "num_appl_by_xstudyidstewards",
    "num_hdp_by_appl",
    "num_hdp_by_xstudyidstewards"
]

# /* General */
# label define yesno 0 "no" 1 "yes"
yesno = {
    0: "no",
    1: "yes"
}


# /* HEAL_01_ImportMerge */
# label define awrep 1 "In reporter only" 2 "In awards only" 3 "In both tables"
awrep = {
    1: "In reporter only",
    2: "In awards only",
    3: "In both tables"
}

# label define sqlmds 1 "In MySQL only" 2 "In MDS only" 3 "In both databases"
sqlmds = {
    1: "In MySQL only",
    2: "In MDS only",
    3: "In both databases"
}

# /* HEAL_04_CTN */
# label define ctn_nums 1 "Project num only in tab Protocol List" 2 "Project num only in tab Complete list CTN Appl_IDs" 3 "Project num in both tabs"
ctn_nums = {
    1: "Project num only in tab Protocol List",
    2: "Project num only in tab Complete list CTN Appl_IDs",
    3: "Project num in both tabs"
}

# /* ----- 1. Import latest MySQL data ----- */
datasets = [
    "reporter",
    "awards",
    "progress_tracker"
]

dfs = {}  # dictionary to hold your dataframes

for name in datasets:
    csv_file = raw / f"{name}_{today}.csv"
    dta_file = raw / f"dta_{name}.csv"

    # read CSV with all columns as strings
    # df = pd.read_csv(csv_file, dtype=str)
    df = pd.read_csv(
    csv_file,
    sep=';',                # Semicolon delimiter
    engine='python',        # Use Python engine for complex parsing
    # quoting=3,              # QUOTE_NONE, avoids treating quotes specially
    encoding='cp1252',
    dtype=str,
    on_bad_lines='warn'   # Skip problematic lines (optional)
    )
    # replace line breaks inside cells with a space
    # and trim whitespace on all string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = (
            df[col]
            .str.replace(r"[\r\n]+", " ", regex=True)
            .str.strip()
        )

    # sort by appl_id (if that column exists)
    if "appl_id" in df.columns:
        df = df.sort_values("appl_id")

    # save to Stata .dta
    df.to_csv(dta_file, index=False)
    print(f"Processed {csv_file} → {dta_file}")

   # store clean DataFrame under a key
    dfs[f"df_{name}"] = df.copy()

   
    
# Load dfs
df_reporter_00 = dfs["df_reporter"]
print("df_reporter_00: " + str(df_reporter_00.shape))
df_awards_00 = dfs["df_awards"]
print("df_awards_00: " + str(df_awards_00.shape))
df_prog_trkr_00 = dfs["df_progress_tracker"]
print("df_prog_trkr_00: " + str(df_prog_trkr_00.shape))


# /* ----- 2. Prepare progress_tracker to merge ----- */

df_prog_trkr_01 = df_prog_trkr_00.copy()
# reorder columns so appl_id comes first
if "appl_id" in df_prog_trkr_01.columns:
    cols = ["appl_id"] + [c for c in df_prog_trkr_01.columns if c != "appl_id"]
    df_prog_trkr_01 = df_prog_trkr_01[cols]

# drop rows where appl_id is blank (empty string) or NaN
df_prog_trkr_01 = df_prog_trkr_01[df_prog_trkr_01["appl_id"].astype(str).str.strip() != ""]

print("df_prog_trkr_01: " + str(df_prog_trkr_01.shape))


# * -- CTN Protocols -- *;

# * Create new CTN variables*;
# * Remove CTN values from appl_id and project_num fields *;

df_prog_trkr_02 = df_prog_trkr_01.copy()

# 1) Create CTN flag: True where project_num starts with "CTN"
df_prog_trkr_02["mds_ctn_flag"] = df_prog_trkr_02["project_num"].str.startswith("CTN", na=False)  # pandas str.startswith for prefix match :contentReference[oaicite:0]{index=0}

# 2) Store CTN project number only for flagged rows
df_prog_trkr_02["mds_ctn_number"] = df_prog_trkr_02["project_num"].where(df_prog_trkr_02["mds_ctn_flag"], None)

# 3) Blank out project_num and appl_id for flagged rows
df_prog_trkr_02.loc[df_prog_trkr_02["mds_ctn_flag"], "project_num"] = ""
df_prog_trkr_02.loc[df_prog_trkr_02["mds_ctn_flag"], "appl_id"] = ""

print("df_prog_trkr_02: " + str(df_prog_trkr_02.shape))



# * -- Project numbers -- *;

# * Split project_num into components needed for xstudy_id *;
# foreach var in project_num {
# replace xproject_num="" if num_dashes>1 /*n=6 changes made*/ 
# 	* Identify and flag bad values of project_num*;
# 	* If an underscore was inserted, remove it and everything that follows it *;
df_prog_trkr_03 = df_prog_trkr_02.copy()

# ————————————————————————
# 1) Create working copy of project_num
# ————————————————————————
df_prog_trkr_03["xproject_num"] = df_prog_trkr_03["project_num"]

# ————————————————————————
# 2) Sieve out characters to count dashes
# (remove everything except dashes)
df_prog_trkr_03["sieved_project_num"] = df_prog_trkr_03["project_num"].str.replace(r"[^-]", "", regex=True)

# Count length of sieved string (trim whitespace)
df_prog_trkr_03["num_dashes"] = df_prog_trkr_03["sieved_project_num"].str.strip().str.len()

# ————————————————————————
# 3) Flag bad project numbers (num_dashes > 1)
df_prog_trkr_03["mds_flag_bad_projnum"] = (df_prog_trkr_03["num_dashes"] > 1).astype(int)
df_prog_trkr_03["mds_bad_projnum"] = df_prog_trkr_03["project_num"].where(df_prog_trkr_03["num_dashes"] > 1, None)

# If more than 1 dash, blank out original in working copy
df_prog_trkr_03.loc[df_prog_trkr_03["num_dashes"] > 1, "xproject_num"] = ""

# ————————————————————————
# 4) Remove underscores and everything after (if any)
df_prog_trkr_03["xproject_num"] = df_prog_trkr_03["xproject_num"].str.replace(r"\_.*", "", regex=True)

# ————————————————————————
# 5) Parse components
# proj_num_spl_ty_code = first character
df_prog_trkr_03["proj_num_spl_ty_code"] = df_prog_trkr_03["xproject_num"].str[0:1]

# proj_num_spl_act_code = next 3 characters
df_prog_trkr_03["proj_num_spl_act_code"] = df_prog_trkr_03["xproject_num"].str[1:4]

# proj_ser_num = next 4 characters
df_prog_trkr_03["proj_ser_num"] = df_prog_trkr_03["xproject_num"].str[4:8]

# ————————————————————————
# 6) Split xproject_num on dashes into parts
split_cols = df_prog_trkr_03["xproject_num"].str.split("-", expand=True)

# drop the first element of the split (like drop xproject_num1 in Stata)
# then take the second element of split as `proj_nm_spl_supp_yr`
# which is at index 1 in pandas because it’s 0-based
df_prog_trkr_03["proj_nm_spl_supp_yr"] = split_cols[1].fillna("")

# suffix code = substr from position 3 onward
df_prog_trkr_03["proj_num_spl_sfx_code"] = df_prog_trkr_03["proj_nm_spl_supp_yr"].str[2:]

# ————————————————————————
# 7) Add `mds_` prefix to match your Stata variable names
df_prog_trkr_03.rename(
    columns={
        "proj_num_spl_ty_code": "mds_proj_num_spl_ty_code",
        "proj_num_spl_act_code": "mds_proj_num_spl_act_code",
        "proj_ser_num": "mds_proj_ser_num",
        "proj_nm_spl_supp_yr": "mds_proj_nm_spl_supp_yr",
        "proj_num_spl_sfx_code": "mds_proj_num_spl_sfx_code",
    },
    inplace=True,
)

# Optional cleanup of intermediate columns
df_prog_trkr_03.drop(columns=["sieved_project_num", "num_dashes"], inplace=True)

print("df_prog_trkr_03: " + str(df_prog_trkr_03.shape))

# * Count number of hdp_ids for a given appl_id *;
# 	/*Note: 2024-10-09: there are only 8 appl_ids that have >1 HDP_ID associated, and the max number of HDP_IDs associated is 3. This excludes CTN records where appl_id==.*/
df_prog_trkr_04 = df_prog_trkr_03.copy()

# 1) count number of hdp_id values per appl_id
num_hdp = (
    df_prog_trkr_04[df_prog_trkr_04["appl_id"].astype(str).str.strip() != ""]  # ignore blank appl_ids
    .groupby("appl_id")["hdp_id"]
    .count()
    .rename("num_hdp_by_appl")
)

# 2) merge that count back into the main DataFrame
df_prog_trkr_04 = df_prog_trkr_04.merge(num_hdp, on="appl_id", how="left")

# 3) rows with blank appl_id → keep count as NaN
mask_blank_appl = df_prog_trkr_04["appl_id"].astype(str).str.strip() == ""
df_prog_trkr_04.loc[mask_blank_appl, "num_hdp_by_appl"] = pd.NA

# 4) rows not found in grouping get 0
df_prog_trkr_04["num_hdp_by_appl"] = df_prog_trkr_04["num_hdp_by_appl"].fillna(0).astype(int)

print("df_prog_trkr_04: " + str(df_prog_trkr_04.shape))


# * Save prepped data *;
df_prog_trkr_05 = df_prog_trkr_04.copy()


# 1) Drop the helper columns
df_prog_trkr_05 = df_prog_trkr_05.drop(columns=["sievedproject_num", "num_dashes", "xproject_num"], errors="ignore")

# 2) Save to Stata .dta
temp = Path(r"\\rtpnfil03\NPTB2\EDC\Migrated\HEAL\Temp")  # replace with your temp folder path
output_file = temp / f"progress_tracker_{today}.csv"

df_prog_trkr_05.to_csv(output_file, index=False)

print(f"Saved prepped data to {output_file}")

print("df_prog_trkr_05: " + str(df_prog_trkr_05.shape))

# /* ----- 3. Merge data ----- */
# * Merge awards reporter *;

# Make local copies so we don’t modify originals
df_reporter_01 = df_reporter_00.copy()
df_awards_01 = df_awards_00.copy()

# —————————————————————————————
# 2) Drop rows where appl_id is blank
df_reporter_01 = df_reporter_01[df_reporter_01["appl_id"].astype(str).str.strip() != ""].copy()
df_awards_01 = df_awards_01[df_awards_01["appl_id"].astype(str).str.strip() != ""].copy()

# —————————————————————————————
# 3) Merge 1:1 on appl_id with indicator
df_nihtables_00 = pd.merge(
    df_reporter_01,
    df_awards_01,
    how="outer",
    on="appl_id",
    indicator="merge_reporter_awards"
)

# —————————————————————————————
# 4) Drop rows where appl_id is blank after merge
df_nihtables_00 = df_nihtables_00[df_nihtables_00["appl_id"].astype(str).str.strip() != ""].copy()

# —————————————————————————————
# 5) Convert indicator text to numeric codes (1/2/3) like Stata
indicator_map = {
    "left_only": 1,   # In reporter only
    "right_only": 2,  # In awards only
    "both": 3         # In both
}

df_nihtables_00["merge_reporter_awards"] = df_nihtables_00["merge_reporter_awards"].map(indicator_map)

# —————————————————————————————
# 6) Save result (CSV or Stata)
df_nihtables_00.to_csv(temp / f"nihtables_{today}.csv", index=False)



print("df_reporter_01: " + str(df_reporter_01.shape))
print("df_awards_01: " + str(df_awards_01.shape))
print("df_nihtables_00: " + str(df_nihtables_00.shape))

# * Merge MDS data (via progress_tracker) *;

# copy so we don’t modify originals
df_nihtables_01 = df_nihtables_00.copy()
df_prog_trkr_06 = df_prog_trkr_05.copy()

# merge 1:m on appl_id with merge indicator
df_dataset_00 = pd.merge(
    df_nihtables_01,
    df_prog_trkr_06,
    on="appl_id",
    how="outer",              # outer keeps all rows from both
    indicator="merge_awards_mds"
)

# map indicator to Stata numeric codes
indicator_map = {
    "left_only": 1,    # In MySQL only
    "right_only": 2,   # In MDS only
    "both": 3          # In both databases
}

df_dataset_00["merge_awards_mds"] = df_dataset_00["merge_awards_mds"].map(indicator_map)

# now df_merged matches Stata
# (you can also add labels or comments for documentation)

# save to Stata or CSV
df_dataset_00.to_csv(temp / f"dataset_{today}.csv", index=False)

print("df_nihtables_01: " + str(df_nihtables_01.shape))
print("df_prog_trkr_06: " + str(df_prog_trkr_06.shape))
print("df_dataset_00: " + str(df_dataset_00.shape))



# /* ----- 4. Clean merged data ----- */
# 	/* Note: subproj_id is not available in the MDS data */ /*n=2 and n=0 real changes made*/

# Make local copies so we don’t modify originals
df_dataset_01 = df_dataset_00.copy()

# Replace blank or whitespace only for proj_ser_num
df_dataset_01.loc[
    df_dataset_01["proj_ser_num"].str.strip().eq(""),
    "proj_ser_num"
] = df_dataset_01.loc[
    df_dataset_01["proj_ser_num"].str.strip().eq(""),
    "mds_proj_ser_num"
]

# Replace blank or whitespace only for proj_num_spl_sfx_code
df_dataset_01.loc[
    df_dataset_01["proj_num_spl_sfx_code"].str.strip().eq(""),
    "proj_num_spl_sfx_code"
] = df_dataset_01.loc[
    df_dataset_01["proj_num_spl_sfx_code"].str.strip().eq(""),
    "mds_proj_num_spl_sfx_code"
]


print("df_dataset_01: " + str(df_dataset_01.shape))

# * Flag supplement awards *;

# Make local copies so we don’t modify originals
df_dataset_02 = df_dataset_01.copy()

# —————————————————————————————
# 1) Flag supplement awards
# —————————————————————————————

# get the second‑to‑last character of proj_num (like Stata substr at -2,1)
df_dataset_02["xsupp_flag"] = df_dataset_02["proj_num"].str[-2:-1]

# flag where that equals "S"
df_dataset_02["supplement_flag"] = (df_dataset_02["xsupp_flag"] == "S").astype(int)

# drop the helper
df_dataset_02.drop(columns=["xsupp_flag"], inplace=True)

# —————————————————————————————
# 2) Convert dates
# —————————————————————————————

# ensure fiscal year numeric
df_dataset_02["fisc_yr"] = pd.to_numeric(df_dataset_02["fisc_yr"], errors="coerce")

# parse the string date columns into real pandas datetime
for var in ["bgt_end", "proj_end_date"]:
    # extract the first 10 characters
    df_dataset_02[f"x{var}"] = df_dataset_02[var].astype(str).str[:10]

    # convert to datetime (pandas datetime)
    df_dataset_02[f"{var}_date"] = pd.to_datetime(df_dataset_02[f"x{var}"], format="%Y-%m-%d", errors="coerce")

    # drop the intermediate
    df_dataset_02.drop(columns=[f"x{var}"], inplace=True)

    # (Optional) reorder columns to come after the original
    cols = list(df_dataset_02.columns)
    if f"{var}_date" in cols:
        cols.remove(f"{var}_date")
        insert_at = cols.index(var) + 1 if var in cols else len(cols)
        cols.insert(insert_at, f"{var}_date")
        df_dataset_02 = df_dataset_02[cols]

    # label for clarity (not necessary in pandas storage)
    df_dataset_02[f"{var}_date"].attrs["label"] = "Python datetime format"

# —————————————————————————————
# 3) Entity type
# —————————————————————————————

# default
df_dataset_02["entity_type"] = "Study"

# override for CTN
df_dataset_02.loc[df_dataset_02["mds_ctn_flag"] == 1, "entity_type"] = "CTN"

# override for bad projnum
df_dataset_02.loc[df_dataset_02["mds_flag_bad_projnum"] == 1, "entity_type"] = "Other"

# —————————————————————————————
# 4) Save result
# —————————————————————————————

df_dataset_02.to_csv(der / f"mysql_{today}.csv", index=False)

print("df_dataset_02: " + str(df_dataset_02.shape))



# /* ----- 2. Generate Research Networks Table ----- */
# /* -------------------------------------------------------------------------------- */
# /* Project: HEAL 																	*/
# /* PI: Kira Bradford, Becky Boyles													*/
# /* Program: HEAL_02_ResNetTable														*/
# /* Programmer: Sabrina McCutchan (CDMS)												*/
# /* Date Created: 2024/05/13															*/
# /* Date Last Updated: 2024/10/09													*/
# /* Description:	This is the Stata program that generates and populates the res_net	*/
# /*				field in the research_networks table in MySQL. 						*/
# /*		1. Import keys 																*/
# /*		2. Create research_networks table											*/
# /*		3. Create data dictionary for research_networks table						*/
# /*		4. Check key contains all values of res_prg									*/
# /*		5. Test MySQL script for generating research_networks table					*/
# /*																					*/
# /* Notes:  																			*/
# /*		- 2024/09/24 this procedure is being migrated to a MySQL Script 			*/
# /*		- 2024/05/21 first run of code to generate research_networks table 			*/

# Helper function to clean a DataFrame
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    # Convert all cells to strings and strip whitespace
    df = df.astype(str).applymap(lambda s: s.replace("\n", " ").strip())
    # Drop empty columns (all values blank or null)
    df = df.replace(r'^\s*$', np.nan, regex=True)
    df = df.dropna(axis=1, how="all")
    # Drop rows where every cell is null
    df = df.dropna(axis=0, how="all")
    return df

# ----- 1. Import keys -----

sheets = ["ref_table", "value_overrides_byappl", "value_overrides_byhdp"]
dfs = {}

for sheet in sheets:
    # import excel sheet into pandas
    
    df = pd.read_excel(
        doc / "HEAL_research_networks_ref_table_for_MySQL.xlsx",
        sheet_name=sheet,
        dtype=str
    )
    df = clean_dataframe(df)
    dfs[sheet] = df

# Save cleaned tables for reuse (optional)
dfs["ref_table"].to_csv( temp / "ref_table.csv", index=False)
dfs["value_overrides_byappl"].to_csv(temp / "value_overrides_byappl.csv", index=False)
dfs["value_overrides_byhdp"].to_csv(temp / "value_overrides_byhdp.csv", index=False)

# ----- Process ref_table -----

ref = dfs["ref_table"]

# Filter out blank res_net
ref = ref[ref["res_net"].str.strip() != ""].copy()

# Uppercase res_net
ref["res_net"] = ref["res_net"].str.upper()

# Only keep relevant columns
ref = ref[["res_prg", "res_net"]]

# Generate lowercase merge key
ref["res_prg_tomerge"] = ref["res_prg"].str.lower().str.replace(r'[^a-z0-9]', '', regex=True)

# Sort
ref = ref.sort_values(by="res_prg_tomerge")

# Save
ref.to_csv(doc / "ref_table.csv", index=False)

# ----- Process value_overrides appl-----

vo = dfs["value_overrides_byappl"]

# Uppercase override
vo["res_net_override"] = vo["res_net"].str.upper()

# Flag
vo["res_net_override_flag"] = 1

# Sort
vo = vo.sort_values(by="appl_id")

# Save
vo.to_csv(doc / "value_overrides_byappl.csv", index=False)

# ----- Process value_overrides hdp-----

vo = dfs["value_overrides_byhdp"]


# Uppercase override
vo["res_net_override"] = vo["res_net"].str.upper()

# Flag
vo["res_net_override_flag"] = 1

# Sort
vo = vo.sort_values(by="hdp_id")

# Save
vo.to_csv(doc / "value_overrides_byhdp.csv", index=False)

# /* ----- 2. Create research_networks table ----- */

# 1) Load the main dataset
df_main = pd.read_csv(temp / f"nihtables_{today}.csv")

# 2) Create merge key (lowercase alpha chars)
df_main['res_prg_tomerge'] = (
    df_main['res_prg']
    .str.lower()
    .str.replace(r'[^a-z0-9]', '', regex=True)
)

# 3) Load reference and overrides
ref = pd.read_csv(doc / "ref_table.csv")          # has columns ['res_prg','res_net','res_prg_tomerge']
vo_appl = pd.read_csv(doc / "value_overrides_byappl.csv")     # from earlier
vo_hdp = pd.read_csv(doc / "value_overrides_byhdp.csv")     # from earlier

# 4) Merge ref_table (many main -> one ref)
df = df_main.merge(
    ref[['res_prg_tomerge', 'res_net']],
    on='res_prg_tomerge',
    how='left',
    indicator='merge_ref_table'
)

# 5) Merge value overrides on appl_id
df = df.merge(
    vo_appl[['appl_id', 'res_net_override', 'res_net_override_flag']],
    on='appl_id',
    how='left'
)

# 6) Apply overrides
df['res_net'] = df.apply(
    lambda r: r['res_net_override'] if r['res_net_override_flag'] == 1 else r['res_net'],
    axis=1
)

# 7) Replace missing override flags (NaN) with 0
df['res_net_override_flag'] = df['res_net_override_flag'].fillna(0).astype(int)

# 8) Keep just needed columns
result = df[['appl_id', 'res_net', 'res_net_override_flag']]

# 9) Save
result.to_csv(der / "research_networks.csv", index=False)


# /* ----- 3. Create data dictionary for research_networks table ----- */
# * Use redcapture command to generate preliminary *;
# * -- Customize to fit Stewards DD template -- *;

# --- 1) Load the reference table and create a “key” ---
ref.to_csv(temp / "ref_table.csv", index=False)  # read Stata .dta
ref = ref[["res_prg"]].drop_duplicates()  # keep unique
ref = ref[ref["res_prg"].notna() & (ref["res_prg"] != "")]  # drop blanks

# --- 2) Load the main dataset and extract unique res_prg ---
main = pd.read_csv(temp / f"dataset_{today}.csv")
main_unique = (
    main[["res_prg"]]
    .drop_duplicates()
    .dropna()
    .loc[lambda df: df["res_prg"] != "0"]
)

# --- 3) Merge to find values not in key ---
merged = main_unique.merge(
    ref,
    on="res_prg",
    how="left",         # left join keeps all main_unique
    indicator=True      # adds _merge column
)

# Keep only those in main but NOT in ref table
missing = merged.loc[merged["_merge"] == "left_only", ["res_prg"]]

# --- 4) Export to CSV ---
missing.to_csv(temp / "missing_res_prg.csv", index=False)

# /* ----- 5. Test MySQL script for generating research_networks table ----- */
# /* Note: This tests a MySQL script to update the research_networks table. It compares the results of the MySQL script and the Stata code. This code block was formerly stroed in the Stata program HEAL_scratch. Initial testing and approval completed 2024/09/23. */

# * Read in MySQL script results *;
# --- 1) Load MySQL results CSV ---
csv_path = backups / "research_networks_2026-03-04.csv"
df_mysql = pd.read_csv(
    csv_path,
    sep=';',                # Semicolon delimiter
    engine='python',        # Use Python engine for complex parsing
    encoding='utf-8'
)

# --- 2) Convert to numeric where appropriate ---
for col in ["res_net_override_flag"]:
    df_mysql[col] = pd.to_numeric(df_mysql[col], errors="coerce")

# --- 3) Uppercase the res_net values ---
df_mysql["res_net"] = df_mysql["res_net"].str.upper()

# --- 4) Show counts of unique res_net values ---
print(df_mysql["res_net"].value_counts(dropna=False))

# --- 5) Sort by appl_id ---
df_mysql = df_mysql.sort_values(by="appl_id")

# --- 6) Optionally save out for later use ---
df_mysql.to_csv(temp / "mysql_resnet.csv", index=False)


# * Report Comparison *;

# --- 1) Load Stata datasets ---
stata_df = pd.read_csv(der / "research_networks.csv")
mysql_df = pd.read_csv(temp / "mysql_resnet.csv")

# --- 2) Compute frequency tables ---
stata_res_net_tab = stata_df["res_net"].value_counts(dropna=False).reset_index()
stata_res_net_tab.columns = ["res_net", "count"]

stata_res_override_tab = stata_df["res_net_override_flag"].value_counts(dropna=False).reset_index()
stata_res_override_tab.columns = ["res_net_override_flag", "count"]

mysql_res_net_tab = mysql_df["res_net"].value_counts(dropna=False).reset_index()
mysql_res_net_tab.columns = ["res_net", "count"]

mysql_res_override_tab = mysql_df["res_net_override_flag"].value_counts(dropna=False).reset_index()
mysql_res_override_tab.columns = ["res_net_override_flag", "count"]

# Generate HTML tables
html_stata_res_net = stata_res_net_tab.to_html(index=False)
html_stata_override = stata_res_override_tab.to_html(index=False)
html_mysql_res_net = mysql_res_net_tab.to_html(index=False)
html_mysql_override = mysql_res_override_tab.to_html(index=False)

# Build HTML page content
html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Research Networks Comparison Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h2 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 80%; margin-bottom: 30px; }}
        th, td {{ border: 1px solid #999; padding: 8px; text-align: left; }}
        th {{ background-color: #eee; }}
    </style>
</head>
<body>
    <h1>Compare MySQL and Stata results for research_networks</h1>

    <h2>Stata res_net</h2>
    {html_stata_res_net}

    <h2>Stata res_net_override_flag</h2>
    {html_stata_override}

    <h2>MySQL res_net</h2>
    {html_mysql_res_net}

    <h2>MySQL res_net_override_flag</h2>
    {html_mysql_override}

</body>
</html>
"""

# Write to an HTML file
with open(temp / "report.html", "w", encoding="utf-8") as f:
    f.write(html_content)

print("HTML report created: report.html")



# * Merge Datasets to Compare *;

# --- 1) Load the MySQL results ---
mysql_df = pd.read_csv(temp / "mysql_resnet.csv")

# Rename columns to avoid conflicts
mysql_df = mysql_df.rename(
    columns={
        "res_net": "mysql_res_net",
        "res_net_override_flag": "mysql_res_net_override"
    }
)

# --- 2) Load the Stata data to compare ---
stata_df = pd.read_csv(der / "research_networks.csv")

# --- 3) Merge on appl_id ---
merged = pd.merge(
    mysql_df,
    stata_df,
    on="appl_id",
    how="inner"  # inner join for matched records only
)

# --- 4) Rename Stata variables to match Stata naming scheme ---
merged = merged.rename(
    columns={
        "res_net": "stata_res_net",
        "res_net_override_flag": "stata_res_net_override"
    }
)

# --- 5) Save merged comparison table ---
merged.to_csv(temp / "compare_res_net.csv", index=False)

# *Override flag discrepancies *;


# Load the merged comparison table we created earlier
compare = pd.read_csv(temp / "compare_res_net.csv")

# Identify rows where the override flag differs
discrepancies = compare.loc[
    compare["stata_res_net_override"] != compare["mysql_res_net_override"]
]

# If you want to just see those discrepant rows:
print(discrepancies)


# *Research network discrepancies *;
# keep if stata_res_net!=mysql_res_net

# Load the merged comparison dataset
compare = pd.read_csv(temp / "compare_res_net.csv")

# Filter where res_net differs
resnet_discrepancies = compare.loc[
    compare["stata_res_net"] != compare["mysql_res_net"]
]

print(f"Number of discrepancies: {len(resnet_discrepancies)}")

# Save just the discrepancy rows to CSV
resnet_discrepancies.to_csv(
    temp / "res_net_discrepancies.csv", index=False
)




# /* ----- 4. Generate CTN crosswalk and outputs ----- */

# /* -------------------------------------------------------------------------------- */
# /* Project: HEAL 																	*/
# /* PI: Kira Bradford																*/
# /* Program: HEAL_09_StudyMetrics													*/
# /* Programmer: Sabrina McCutchan (CDMS)												*/
# /* Date Created: 2024/06/23															*/
# /* Date Last Updated: 2025/12/11													*/
# /* Description:	This program produces a report of HDE study metrics.				*/
# /*		1. Number of NIH awards in MySQL											*/
# /*		2. Number of studies with VLMD on Platform									*/
# /*		3. Number of studies with VLMD available in HSS								*/
# /*		4. Number of studies who've submitted CDE usage								*/
# /*		5. HEAL studies by data sharing intention									*/
# /*		6. Number of studies registered 											*/
# /*		7. Number of studies submitted SLMD											*/
# /*		8. Number of studies selected repo											*/
# /*		9. Count of how many studies selected each repo								*/
# /*		10. Number of studies with data linked on Platform							*/
# /*		11. Other: MySQL entities and awards										*/
# /*																					*/
# /* Notes:  																			*/
# /*		- Metrics should be calculated using only Platform data.					*/
# /*		- In June 2024, a "6-Month Report" was generated to measure Get the Data	*/
# /*		  (GtD) progress. This contained metrics which resemble but predate the		*/
# /*		  formally agreed-upon study metrics. Thus, this program supersedes the 	*/
# /*		  previous program HEAL_6Mo_Report.do; the latter has been archived.		*/
# /*																					*/
# /* Version changes																	*/
# /*		- 2025/12/11 - added count of how many studies selected each repo (#9)		*/
# /*		- 2025/06/16 - categories from Data Dictionary Tracker monday board changed */
# /*		- 2025/04/25 - Met w/ Kathy and Hina and removed conditional exclusion of	*/
# /*		  drop if gen3_data_availability=="not_available" from registration and		*/
# /*		  SLMD completion metrics.													*/
# /*		- 2024/12/13 - Code updated after changing MySQL's progress_tracker table	*/
# /*		  to include more fields and getting specification from the Platform on how */
# /*		  they calculate metrics. See "Study Tracking in Platform and MySQL_Running */
# /*		  Agenda", section "Revised HDE Metrics, Definitions and Data Sources". 	*/
# /*		- v1 - HEAL_6Mo_Report.do, now archived. June 2024. 						*/	
# /*																					*/
# /* -------------------------------------------------------------------------------- */

# /* ----- 0. Prepare standard dataset for metrics report ----- */
df = pd.read_csv(f"{der}/mysql_{today}.csv")

df = df[df["merge_awards_mds"] != 1]

df = df[df["guid_type"].isin([
    "discovery_metadata",
    "unregistered_discovery_metadata"
])]

metrics = df.copy()
metrics.to_csv(f"{temp}/metrics_{today}.csv", index=False)



# /* ----- 1. Number of HEAL Studies ----- */
heal_studies = metrics[["hdp_id","entity_type"]].copy()
heal_studies["HEAL_studies"] = range(1, len(heal_studies)+1)

print("1. Number of HEAL Studies", heal_studies["HEAL_studies"].max())
print("\nBreakdown by entity_type")
print(heal_studies["entity_type"].value_counts())
freq_entity_type = (
    heal_studies['entity_type']
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_entity_type.columns = ['entity_type', 'count']

from openpyxl import load_workbook  

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_template.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

ws["B2"] = today
ws["B5"] = heal_studies["HEAL_studies"].max()
start_row = 9  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_entity_type.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# ----- 2. Studies with VLMD on Platform -----
df2 = metrics.copy()

df2["num_data_dictionaries"] = pd.to_numeric(
    df2["num_data_dictionaries"], errors="coerce"
)

df2 = df2[df2["num_data_dictionaries"] > 0]

df2 = df2[["hdp_id","num_data_dictionaries"]]
df2["vlmd_available_platform"] = range(1, len(df2)+1)

print("Studies with VLMD on platform:", df2["vlmd_available_platform"].max())

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  
ws["B17"] = df2["vlmd_available_platform"].max()
wb.save(f"{out}/StudyMetrics_{today}.xlsx")




# /* ----- 3. Number of studies with VLMD available in HSS ----- */

# /* ----- 4. Number of studies who've submitted CDE usage----- */
df4 = metrics.copy()

df4["num_common_data_elements"] = pd.to_numeric(
    df4["num_common_data_elements"], errors="coerce"
)

df4 = df4[df4["num_common_data_elements"] > 0]

df4 = df4[["hdp_id","num_common_data_elements"]]
df4["cdes"] = range(1, len(df4)+1)

print("Studies reporting CDE usage:", df4["cdes"].max())

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  
ws["B26"] = df4["cdes"].max()
wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# ----- 5. HEAL studies by data sharing intention -----
print(metrics["gen3_data_availability"].value_counts(dropna=False))

freq_data_availability = (
    metrics["gen3_data_availability"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_data_availability.columns = ["gen3_data_availability", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 32  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_data_availability.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# ----- 6. Number of studies registered -----
print(metrics["is_registered"].value_counts(dropna=False))

freq_is_registered = (
    metrics["is_registered"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_is_registered.columns = ["is_registered", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 43  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_is_registered.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# ----- 7. Studies submitted SLMD -----
df7 = metrics.copy()

df7["overall_percent_complete"] = pd.to_numeric(
    df7["overall_percent_complete"], errors="coerce"
)

df7["slmd"] = np.where(
    (df7["overall_percent_complete"] >= 50),
    1,
    0
)

print(df7["slmd"].value_counts(dropna=False))

freq_slmd = (
    df7["slmd"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_slmd.columns = ["slmd", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 51  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_slmd.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")




# ----- 8. Studies selecting repo -----
df8 = metrics.copy()

df8 = df8[df8["gen3_data_availability"] != "not_available"]

df8["has_repo"] = np.where(
    df8["repository_name"].str.strip() == "",
    0,
    1
)

print(df8["has_repo"].value_counts(dropna=False))

freq_has_repo = (
    df8["has_repo"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_has_repo.columns = ["has_repo", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 59  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_has_repo.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# ----- 9. Number of studies selecting each repo -----
df9 = metrics.copy()

df9 = df9[df9["gen3_data_availability"] != "not_available"]

df9 = df9[df9["repository_name"].str.strip() != ""]

repo_counts = df9["repository_name"].value_counts()

print("\nStudies selecting each repository")
print(repo_counts)

freq_repo_counts = (
    df9["repository_name"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_repo_counts.columns = ["repository_name", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Repository Names"]  # Replace with your sheet name  

start_row = 2  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_repo_counts.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# ----- 10. Studies with data linked on platform -----
df10 = metrics.copy()

df10 = df10[df10["gen3_data_availability"] != "not_available"]

print(df10["data_linked_on_platform"].value_counts(dropna=False))

freq_link_platf = (
    df10["data_linked_on_platform"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_link_platf.columns = ["data_linked_on_platform", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 105  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_link_platf.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")





# /* ----- 11. Other: MySQL entities and awards ----- */
# asdoc, text(--------------11. Other: MySQL entities and awards--------------) fs(14), save($qc/StudyMetrics_$today.doc) append label

# * Number of CTN protocols *;
# use "$der/mysql_$today.dta", clear
# keep if mds_ctn_flag==1
# gen ctns=_n
# label var ctns "Number of CTN Protocols"
# asdoc sum ctns, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of CTN Protocols is determined by how many HDP IDs on Platform are records for a CTN Protocol.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label

# * Number of MySQL studies *;
# use "$der/study_lookup_table.dta", clear
# destring xstudy_id, replace
# label var xstudy_id "Number of MySQL studies"
# asdoc sum xstudy_id, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of distinct studies tracked in the MySQL DB, where 'study' is defined by the HEAL Stewards. The number reported above is a count of unique values of xstudy_id.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label

# * Number of awards in MySQL *;
# use "$der/mysql_$today.dta", clear
# drop if merge_awards_mds==2
# keep appl_id
# sort appl_id
# drop if appl_id==""
# duplicates drop
# gen appls=_n
# label var appls "Number of MySQL awards"
# asdoc sum appls, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of awards tracked in the MySQL DB, where an award is defined as a unique NIH appl_id.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label


# * Awards in MySQL by entity type *;
# use "$der/mysql_$today.dta", clear
# label var entity_type "Entity Type"
# drop if merge_awards_mds==2
# keep appl_id entity_type
# sort appl_id
# duplicates drop
# merge 1:1 appl_id using "$raw/research_networks_$today.dta"
# replace entity_type="CTN Protocol" if res_net=="CTN"
# asdoc tab entity_type, title(appl_ids by entity type) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(This tabulation shows the number of awards belonging to each type of entity. Here, the number for CTN Protocol indicates the total number of appl_ids associated with any CTN Protocol numbers. The 'Other' entity type indicates 6 awards that do not have project serial numbers and appear to be contracts or other agreements - these 6 are listed out in the regular QC Report. Every other appl_id in the MySQL DB belongs to a Study, where 'study' is defined by the Stewards.) save($qc/StudyMetrics_$today.doc) append label




# ----- 11. Other metrics -----

mysql = pd.read_csv(f"{der}/mysql_{today}.csv")

# CTN protocols
ctn = mysql[mysql["mds_ctn_flag"] == 1]
print("Number of CTN Protocols:", len(ctn))

# MySQL studies
# read CSV with all columns as strings
studies = pd.read_csv(
    f"{der}/study_lookup_table.csv",
    sep=';',                # Semicolon delimiter
    engine='python',        # Use Python engine for complex parsing
    # quoting=3,              # QUOTE_NONE, avoids treating quotes specially
    encoding='cp1252',
    dtype=str,
    on_bad_lines='warn'   # Skip problematic lines (optional)
    )
    # replace line breaks inside cells with a space
    # and trim whitespace on all string columns
for col in studies.select_dtypes(include="object").columns:
        studies[col] = (
            studies[col]
            .str.replace(r"[\r\n]+", " ", regex=True)
            .str.strip()
        )



studies["xstudy_id"] = pd.to_numeric(studies["xstudy_id"], errors="coerce")

print("Number of MySQL studies:", studies["xstudy_id"].nunique())

# MySQL awards
awards = mysql[mysql["merge_awards_mds"] != 2]
awards = awards[awards["appl_id"] != ""]
awards = awards["appl_id"].drop_duplicates()

print("Number of MySQL awards:", len(awards))

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

ws["B112"] = len(ctn)
ws["B114"] = studies["xstudy_id"].nunique()
ws["B116"] = len(awards)




start_row = 119  # Excel row number (A2 = row 5)  
start_col = 1  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_entity_type.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# Write MySQL Tables to Worksheets
wb_filepath = f"{out}/StudyMetrics_{today}.xlsx"

# Use ExcelWriter in append mode ('a')
# Pandas will automatically handle loading the existing workbook
with pd.ExcelWriter(wb_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    df_awards_00.to_excel(writer, sheet_name='MySQL_Awards', index=False)
    
with pd.ExcelWriter(wb_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    df_prog_trkr_00.to_excel(writer, sheet_name='MySQL_Prog_Tracker', index=False)


# This regex matches common illegal characters in Excel (control characters)
# except for standard newlines (\n), carriage returns (\r), and tabs (\t).
def remove_excel_illegal_chars(val):
    if isinstance(val, str):
        return re.sub(r'[\000-\010\013\014\016-\037]', '', val)
    return val

# Apply to the dataframe causing the error
df_reporter_00 = df_reporter_00.map(remove_excel_illegal_chars)

# Now try writing to Excel again
with pd.ExcelWriter(wb_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_reporter_00.to_excel(writer, sheet_name='MySQL_Reporter', index=False)
    
    
with pd.ExcelWriter(wb_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    # 3. Write your DataFrame directly
    df_mysql.to_excel(writer, sheet_name='MySQL_Research_Networks', index=False)
    
    
    

