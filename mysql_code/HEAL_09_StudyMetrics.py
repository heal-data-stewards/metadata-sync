# /* -------------------------------------------------------------------------------- */
# /* Project: HEAL 																	*/
# /* PI: Kira Bradford																*/
# /* Program: HEAL_09_StudyMetrics													*/
# /* Programmer: Sabrina McCutchan (CDMS)												*/
# /* Date Created: 2024/06/23															*/
# /* Date Last Updated: 2025/12/11													*/
# /* Description:	This program produces a report of HDE study metrics.				*/
# /*		1. Number of NIH awards in MySQL											*/
# /*		2. Number of studies with VLMD on Platform									*/
# /*		3. Number of studies with VLMD available in HSS								*/
# /*		4. Number of studies who've submitted CDE usage								*/
# /*		5. HEAL studies by data sharing intention									*/
# /*		6. Number of studies registered 											*/
# /*		7. Number of studies submitted SLMD											*/
# /*		8. Number of studies selected repo											*/
# /*		9. Count of how many studies selected each repo								*/
# /*		10. Number of studies with data linked on Platform							*/
# /*		11. Other: MySQL entities and awards										*/
# /*																					*/
# /* Notes:  																			*/
# /*		- Metrics should be calculated using only Platform data.					*/
# /*		- In June 2024, a "6-Month Report" was generated to measure Get the Data	*/
# /*		  (GtD) progress. This contained metrics which resemble but predate the		*/
# /*		  formally agreed-upon study metrics. Thus, this program supersedes the 	*/
# /*		  previous program HEAL_6Mo_Report.do; the latter has been archived.		*/
# /*																					*/
# /* Version changes																	*/
# /*		- 2025/12/11 - added count of how many studies selected each repo (#9)		*/
# /*		- 2025/06/16 - categories from Data Dictionary Tracker monday board changed */
# /*		- 2025/04/25 - Met w/ Kathy and Hina and removed conditional exclusion of	*/
# /*		  drop if gen3_data_availability=="not_available" from registration and		*/
# /*		  SLMD completion metrics.													*/
# /*		- 2024/12/13 - Code updated after changing MySQL's progress_tracker table	*/
# /*		  to include more fields and getting specification from the Platform on how */
# /*		  they calculate metrics. See "Study Tracking in Platform and MySQL_Running */
# /*		  Agenda", section "Revised HDE Metrics, Definitions and Data Sources". 	*/
# /*		- v1 - HEAL_6Mo_Report.do, now archived. June 2024. 						*/	
# /*																					*/
# /* -------------------------------------------------------------------------------- */

# CURRENTLY IMPORTS CSV files exported from MySQL Heal Studies database
# Requires these _today.csv files to exist:

# ----- Boiler Plate Code -----------------------------------------------------*/


# Import Python Modules Here
import os
import pandas as pd
from datetime import date
import subprocess
import sys
from pathlib import Path
import re
import numpy as np
import openpyxl
from openpyxl import load_workbook  
 
# ----- SET MACROS -----*/

# ----- 1. Dates ----- */
today = "2026-04-24"
# today = date.today().strftime("%Y-%m-%d")
print(today)

# ----- 2. Filepaths ----- */
# dir = Path(r"/rtpnfil03/rtpnfil03_vol4/NPTB2/EDC/Migrated/HEAL") #UNIX
dir = Path(r"\\rtpnfil03\NPTB2\EDC\Migrated\HEAL") #WINDOwS
raw = dir / "Extracts"
der = dir / "Derived"
prog = dir / "Programs"
doc = dir / "Documentation"
temp = dir / "temp"
out = dir / "Output"
qc = out / "QC"
backups = dir / "Backups"

# Logging
log_path = os.path.join(out, f"HEAL_09_StudyMetrics_{today}_log.txt")
with open(log_path, 'w') as f:
    pass  # 'w' mode truncates existing file or creates new blank file
# open(f"{out}/StudyMetrics_{today}_log.txt", 'w').close() #Clears Log before running

def log_out(message):
    with open(f"{out}/HEAL_09_StudyMetrics_{today}_log.txt", 'a') as f:
        print(message, file=f)

log_out(f"HEAL_09_StudyMetrics Log Run Date: {today}")

# ----- END Boiler Plate Code -----------------------------------------------------*/
 



# HEAL_09_StudyMetrics

# guid_type hx table
log_out(f"HEAL_09_StudyMetrics")

# /* ----- 0. Prepare standard dataset for metrics report ----- */


df = pd.read_csv(f"{der}/mysql_{today}.csv")
log_out(f"df: Using mysql_today.csv' {df.shape}")

# df = df[df["merge_awards_mds"] != 1]
df = df[df["merge_awards_mds"].astype(str).str[0] != '1']  #now that merge_awards_mds is labeled for readability
log_out(f"df: exclude where merge_awards_mds = 1' {df.shape}")

df = df[df["guid_type"].isin([
    "discovery_metadata",
    "unregistered_discovery_metadata"
])]
log_out(f"df: keep where guid_type in (discovery_metadata,unregistered_discovery_metadata)' {df.shape}")

metrics = df.copy()
metrics.to_csv(f"{temp}/metrics_{today}.csv", index=False)
log_out(f"metrics: renamed copy of df and saved to metrics_today.csv' {metrics.shape}")



# /* ----- 1. Number of HEAL Studies C5----- */
heal_studies = metrics[["hdp_id","entity_type"]].copy()
heal_studies["HEAL_studies"] = range(1, len(heal_studies)+1)

print("1. Number of HEAL Studies", heal_studies["HEAL_studies"].max())
print("\nBreakdown by entity_type")
print(heal_studies["entity_type"].value_counts())
freq_entity_type = (
    heal_studies['entity_type']
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_entity_type.columns = ['entity_type', 'count']

from openpyxl import load_workbook  

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_template_v2.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

ws["C2"] = today

# 1. Number of HEAL Studies C5, C9, C10, C11
ws["C5"] = heal_studies["HEAL_studies"].max()

entity_type = "Study"
ws["C9"] = heal_studies[heal_studies["entity_type"] == entity_type]["HEAL_studies"].max()

entity_type = "CTN"
ws["C10"] = heal_studies[heal_studies["entity_type"] == entity_type]["HEAL_studies"].max()

entity_type = "Other"
ws["C11"] = heal_studies[heal_studies["entity_type"] == entity_type]["HEAL_studies"].max()

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# 2. Number of studies registered B18-----
print(metrics["is_registered"].value_counts(dropna=False))

freq_is_registered = (
    metrics["is_registered"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_is_registered.columns = ["is_registered", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 18  # Excel row number (A2 = row 5)  
start_col = 2  # Excel column number (A = column 1)

df_values = freq_is_registered.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# 3. Number of studies who've submitted CDE usage C25----- */
df3 = metrics.copy()

df3["num_common_data_elements"] = pd.to_numeric(
    df3["num_common_data_elements"], errors="coerce"
)

df3 = df3[df3["num_common_data_elements"] > 0]

df3 = df3[["hdp_id","num_common_data_elements"]]
df3["cdes"] = range(1, len(df3)+1)

print("Studies reporting CDE usage:", df3["cdes"].max())

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  
ws["C25"] = df3["cdes"].max()
wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# ----- 4. Studies submitted SLMD B31-----
df4 = metrics.copy()

df4["overall_percent_complete"] = pd.to_numeric(
    df4["overall_percent_complete"], errors="coerce"
)

df4["slmd"] = np.where(
    (df4["overall_percent_complete"] >= 50),
    1,
    0
)

print(df4["slmd"].value_counts(dropna=False))

freq_slmd = (
    df4["slmd"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_slmd.columns = ["slmd", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 31  # Excel row number
start_col = 2  # Excel column number 

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_slmd.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# 5. HEAL studies by data sharing intention B39-----
print(metrics["gen3_data_availability"].value_counts(dropna=False))

freq_data_availability = (
    metrics["gen3_data_availability"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_data_availability.columns = ["gen3_data_availability", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 39  # Excel row number (A2 = row 5)  
start_col = 2  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_data_availability.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# 6. Studies selecting repo B50 USE repository_selected-----
df6 = metrics.copy()

df6 = df6[df6["gen3_data_availability"] != "not_available"] #producing and sharing data


# df6["has_repo"] = np.where(
#     df6["repository_name"].str.strip() == "",
#     0,
#     1
# )

# print(df6["has_repo"].value_counts(dropna=False))

# freq_has_repo = (
#     df6["has_repo"]
#     .value_counts(dropna=False)   # include missing like SAS
#     .reset_index()
# )

# freq_has_repo.columns = ["has_repo", 'count']

# Alt version of has_repo
freq_has_repo = (
    df6["repository_selected"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_has_repo.columns = ["repository_selected", 'count']
print(df6["repository_selected"].value_counts(dropna=False))


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 50  # Excel row number (A2 = row 5)  
start_col = 2  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_has_repo.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")



# 8. Number of studies selecting each repo 8.Repository Names Worksheet-----
df8 = metrics.copy()

df8 = df8[df8["gen3_data_availability"] != "not_available"]

df8 = df8[df8["repository_name"].str.strip() != ""]

repo_counts = df8["repository_name"].value_counts()

print("\nStudies selecting each repository")
print(repo_counts)

freq_repo_counts = (
    df8["repository_name"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_repo_counts.columns = ["repository_name", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["8. Repository Names"]  # Replace with your sheet name  

start_row = 3  # Excel row number 
start_col = 1  # Excel column number 

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_repo_counts.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")




# 9. Studies with VLMD on Platform C59-----
df9 = metrics.copy()

df9["num_data_dictionaries"] = pd.to_numeric(
    df9["num_data_dictionaries"], errors="coerce"
)

df9 = df9[df9["num_data_dictionaries"] > 0]

df9 = df9[["hdp_id","num_data_dictionaries"]]
df9["vlmd_available_platform"] = range(1, len(df9)+1)

print("Studies with VLMD on platform:", df9["vlmd_available_platform"].max())

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  
ws["C59"] = df9["vlmd_available_platform"].max()
wb.save(f"{out}/StudyMetrics_{today}.xlsx")





# 10. Number of studies with VLMD available in HSS  "See Vicki's report----- */




# 11. Studies with data linked on platform B69-----
df11 = metrics.copy()

df11 = df11[df11["gen3_data_availability"] != "not_available"]

print(df11["data_linked_on_platform"].value_counts(dropna=False))

freq_link_platf = (
    df11["data_linked_on_platform"]
    .value_counts(dropna=False)   # include missing like SAS
    .reset_index()
)

freq_link_platf.columns = ["data_linked_on_platform", 'count']


# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

start_row = 69  # Excel row number (A2 = row 5)  
start_col = 2  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

df_values = freq_link_platf.values  

# Iterate over rows and columns to write values  
for row_idx, row_data in enumerate(df_values):  
    for col_idx, value in enumerate(row_data):  
        # Calculate Excel cell coordinates (1-based)  
        excel_row = start_row + row_idx  
        excel_col = start_col + col_idx  
 
        # Write only the value (preserves formatting)  
        ws.cell(row=excel_row, column=excel_col).value = value

wb.save(f"{out}/StudyMetrics_{today}.xlsx")





# /* ----- 11. Other: MySQL entities and awards ----- */
# asdoc, text(--------------11. Other: MySQL entities and awards--------------) fs(14), save($qc/StudyMetrics_$today.doc) append label

# * Number of CTN protocols *;
# use "$der/mysql_$today.dta", clear
# keep if mds_ctn_flag==1
# gen ctns=_n
# label var ctns "Number of CTN Protocols"
# asdoc sum ctns, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of CTN Protocols is determined by how many HDP IDs on Platform are records for a CTN Protocol.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label

# * Number of MySQL studies *;
# use "$der/study_lookup_table.dta", clear
# destring xstudy_id, replace
# label var xstudy_id "Number of MySQL studies"
# asdoc sum xstudy_id, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of distinct studies tracked in the MySQL DB, where 'study' is defined by the HEAL Stewards. The number reported above is a count of unique values of xstudy_id.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label

# * Number of awards in MySQL *;
# use "$der/mysql_$today.dta", clear
# drop if merge_awards_mds==2
# keep appl_id
# sort appl_id
# drop if appl_id==""
# duplicates drop
# gen appls=_n
# label var appls "Number of MySQL awards"
# asdoc sum appls, statistics(max) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(The number of awards tracked in the MySQL DB, where an award is defined as a unique NIH appl_id.) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text( ) save($qc/StudyMetrics_$today.doc) append label


# * Awards in MySQL by entity type *;
# use "$der/mysql_$today.dta", clear
# label var entity_type "Entity Type"
# drop if merge_awards_mds==2
# keep appl_id entity_type
# sort appl_id
# duplicates drop
# merge 1:1 appl_id using "$raw/research_networks_$today.dta"
# replace entity_type="CTN Protocol" if res_net=="CTN"
# asdoc tab entity_type, title(appl_ids by entity type) save($qc/StudyMetrics_$today.doc) append label
# asdoc, text(This tabulation shows the number of awards belonging to each type of entity. Here, the number for CTN Protocol indicates the total number of appl_ids associated with any CTN Protocol numbers. The 'Other' entity type indicates 6 awards that do not have project serial numbers and appear to be contracts or other agreements - these 6 are listed out in the regular QC Report. Every other appl_id in the MySQL DB belongs to a Study, where 'study' is defined by the Stewards.) save($qc/StudyMetrics_$today.doc) append label




# 12. Other metrics C76 C78 C80 B83-----

mysql = pd.read_csv(f"{der}/mysql_{today}.csv")

# CTN protocols
ctn = mysql[mysql["mds_ctn_flag"] == 1]
print("Number of CTN Protocols:", len(ctn))

# MySQL studies
# read CSV with all columns as strings
studies = pd.read_csv(
    f"{der}/study_lookup_table.csv",
    sep=';',                # Semicolon delimiter
    engine='python',        # Use Python engine for complex parsing
    # quoting=3,              # QUOTE_NONE, avoids treating quotes specially
    encoding='cp1252',
    dtype=str,
    on_bad_lines='warn'   # Skip problematic lines (optional)
    )
    # replace line breaks inside cells with a space
    # and trim whitespace on all string columns
for col in studies.select_dtypes(include="object").columns:
        studies[col] = (
            studies[col]
            .str.replace(r"[\r\n]+", " ", regex=True)
            .str.strip()
        )



studies["xstudy_id"] = pd.to_numeric(studies["xstudy_id"], errors="coerce")

print("Number of MySQL studies:", studies["xstudy_id"].nunique())

# MySQL awards
awards = mysql[mysql["merge_awards_mds"].astype(str).str[0] != '2']
awards = awards[awards["appl_id"] != ""]
awards = awards["appl_id"].drop_duplicates()

print("Number of MySQL awards:", len(awards))

# Load the template (read_only=False to allow writing)  
wb = load_workbook(f"{out}/StudyMetrics_{today}.xlsx")  
 
# Select the worksheet to write to (e.g., "Report")  
ws = wb["Metrics"]  # Replace with your sheet name  

ws["C76"] = len(ctn)
ws["C78"] = studies["xstudy_id"].nunique()
ws["C80"] = len(awards)




# start_row = 83  # Excel row number (A2 = row 5)  
# start_col = 2  # Excel column number (A = column 1)

# Extract DataFrame values as a numpy array  
# freq_entity_type.columns = ['entity_type', 'count']

# df_values = freq_entity_type.values  

# Iterate over rows and columns to write values  
# for row_idx, row_data in enumerate(df_values):  
#     for col_idx, value in enumerate(row_data):  
#         # Calculate Excel cell coordinates (1-based)  
#         excel_row = start_row + row_idx  
#         excel_col = start_col + col_idx  
 
#         # Write only the value (preserves formatting)  
#         ws.cell(row=excel_row, column=excel_col).value = value

# wb.save(f"{out}/StudyMetrics_{today}.xlsx")


# Write MySQL Tables to Worksheets
# wb_filepath = f"{out}/StudyMetrics_{today}.xlsx"

# with pd.ExcelWriter(wb_filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#     # 3. Write your DataFrame directly
#     mysql_today_noabs.to_excel(writer, sheet_name='Derived_mysql_today', index=False)

    
# END HEAL_09_StudyMetrics
    
    

